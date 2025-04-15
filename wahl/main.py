import os
import random

import openpyxl as ex
from openpyxl import Workbook

path = "wahl/test_tabelle.xlsx"
company_path = "wahl/company_names.xlsx"
n = 1
wb = ex.load_workbook(path)
sheet = wb.active

row = sheet.max_row
column = sheet.max_column

company_wb = ex.load_workbook(company_path)
company_sheet = company_wb.active

company_row = company_sheet.max_row
company_column = company_sheet.max_column

STUDENTS = []
COMPANIES = []
random.seed(os.urandom(8))


class Student:

    def __init__(self, name, group, choices):
        self.name = name
        self.group = group
        self.choices = choices
        self.wish_num = 0
        self.course_num = 0
        self.courses = []


class Company:

    def __init__(self, name):
        self.name = name
        self.students = [[],[],[]]
        self.max_space = 10


def read_data():
    for i in range(2, row + 1):
        current_student = []
        for k in range(2, column + 1):
            cell = sheet.cell(row=i, column=k)
            current_student.append(cell.value)
        p = Student(current_student[0], current_student[1], [current_student[2], current_student[3], current_student[4], current_student[5], current_student[6], current_student[7]])
        STUDENTS.append(p)

    for i in range(2, company_row + 1):
        current_company = []
        for k in range(1, company_column + 1):
            cell = company_sheet.cell(row=i, column=k)
            current_company.append(cell.value)
        p = Company(current_company[0])
        COMPANIES.append(p)


def add_student(wish_num, curr_student):
    comp_name = curr_student.choices[wish_num]
    num = get_num(comp_name) - 1

    if len(COMPANIES[num].students[curr_student.course_num]) == COMPANIES[num].max_space:
        if curr_student.wish_num < 5:
            curr_student.wish_num += 1
            add_student(curr_student.wish_num, curr_student)
        else:
            append_course(curr_student, "error2")
    else:

        
        if curr_student.course_num < 3:
            COMPANIES[num].students[curr_student.course_num].append(curr_student.name)
            curr_student.courses.append(comp_name)
            curr_student.course_num = curr_student.course_num + 1
        

def get_num(name):
    global n
    print(n)
    print(name)
    x = name.split(".", 1)
    n += 1
    return int(x[0])
    

def main():
    for i in range(6):
        random.shuffle(STUDENTS)
        for curr_student in STUDENTS:
            if len(curr_student.courses) < 3 and i + curr_student.wish_num < 6:
                add_student(i + curr_student.wish_num, curr_student)
            elif len(curr_student.courses) < 3 and i + curr_student.wish_num >= 6:
                append_course(curr_student, "error")


def append_course(student, course):
    student.courses.append(course)
    student.course_num += 1


def write_data1():
    wb1 = Workbook()
    sheet1 = wb1.active

    sheet1["A1"].value = "Name"
    sheet1["B1"].value = "Klasse"
    sheet1["C1"].value = "Block 1"
    sheet1["D1"].value = "Block 2"
    sheet1["E1"].value = "Block 3"

    for i in range(2, len(STUDENTS) + 2):
        curr = STUDENTS[i - 2]
        l = [curr.name, curr.group, curr.courses[0], curr.courses[1], curr.courses[2]]

        for k in range(1, 6):
            sheet1.cell(row=i, column=k).value = l[k - 1]

    wb1.save(filename="liste.xlsx")


def write_data2(it, name):
    wb1 = Workbook()
    sheet1 = wb1.active
    
    for i in range(1, len(COMPANIES) + 1):
        print(len(COMPANIES[i - 1].students[it]))
        sheet1.cell(row=1, column=i).value = COMPANIES[i - 1].name
        for k in range(2, len(COMPANIES[i - 1].students[it]) + 2):
            print(COMPANIES[i - 1].students[it] [k - 2])
            sheet1.cell(row=k, column=i).value = COMPANIES[i - 1].students[it] [k - 2]
    
    wb1.save(filename = name)

def write_data3():
    wb1 = Workbook()
    sheet1 = wb1.active

read_data()
main()
write_data1()
write_data2(0,"tb1.xlsx")
write_data2(1,"tb2.xlsx")
write_data2(2,"tb3.xlsx")

